import requests
import re
import openpyxl
import csv
import os
from openpyxl import Workbook
from config import *
from slackmanger import *
from openpyxl.utils import get_column_letter
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText

class RealEstateScraper:
    def __init__(self, cookies, headers, apt_list_file):
        self.cookies = cookies
        self.headers = headers
        self.apt_list_file = apt_list_file
        self.apt_data = self.load_apt_data()
        self.email_sender = EMAIL_SENDER
        self.email_smtp_pw = EMAIL_SMTP_PW
    
    def load_apt_data(self):
        apt_data = {}
        with open(self.apt_list_file, mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file, delimiter=',')
            for row in reader:
                if row['enable'] == 'X':
                    continue
                apt_name = row['아파트 이름']
                apt_idx = row['apt index']
                area_sizes = row['평형'].split(',')  # 여러 평형을 쉼표로 구분
                apt_data[apt_idx] = {'apt_name': apt_name, 'area_sizes': [area.strip() for area in area_sizes]}
        return apt_data

    def parse_price(self, price_str):
        # 공백 제거
        price_str = price_str.replace(" ", "")
        # 정규식을 통해 금액 분석
        match = re.match(r"(\d+)(억)(\s*(\d{1,3}(,\d{3})*))?", price_str)
        if match:
            billion = int(match.group(1))  # 억 단위 추출
            million = match.group(4)  # 만 단위 추출 (콤마 포함)
            price = billion * 100000000  # 억 단위 변환
            if million:
                # 콤마 제거 후 정수로 변환
                million = int(million.replace(",", ""))
                price += million * 10000  # 만 단위 변환
            return price
        return 0

    def remove_duplicate_lists(self, data):
        unique_data = []
        for item in data:
            # 첫 번째 항목이 이미 있는지 확인
            if not any(
                all(item[i] == existing_item[i] for i in [0, 3, 4, 6, 7, 9]) for existing_item in unique_data
            ):
                unique_data.append(item)
        return unique_data

    def fetch_data(self, apt_idx):
        base_url = f'https://new.land.naver.com/api/articles/complex/{apt_idx}?realEstateType=APT%3AABYG%3AJGC%3APRE&tradeType=B2%3AA1&tag=%3A%3A%3A%3A%3A%3A%3A%3A&rentPriceMin=0&rentPriceMax=900000000&priceMin=0&priceMax=900000000&areaMin=0&areaMax=900000000&oldBuildYears&recentlyBuildYears&minHouseHoldCount&maxHouseHoldCount&showArticle=false&sameAddressGroup=true&minMaintenanceCost&maxMaintenanceCost&priceType=RETAIL&directions=&complexNo=120960&buildingNos=&areaNos=14%3A1%3A3%3A2%3A4%3A5%3A7%3A8%3A6%3A9%3A10%3A11%3A15%3A12&type=list&order=rank'

        all_articles = []
        page = 1
        while True:
            response = requests.get(f"{base_url}&page={page}", cookies=self.cookies, headers=self.headers)
            if response.status_code != 200:
                break
            
            data = response.json()
            articles = data.get('articleList', [])

            for article in articles:
                str_area2 = str(article.get('area2'))
                if article.get('tradeTypeName') == '매매' and str_area2 in self.apt_data[apt_idx]['area_sizes']:
                    all_articles.append([
                        article.get('articleName', ''),
                        article.get('realEstateTypeName', ''),
                        article.get('tradeTypeName', ''),
                        article.get('dealOrWarrantPrc', ''),
                        article.get('areaName', ''),
                        article.get('area2', ''),
                        article.get('floorInfo', ''),
                        article.get('direction', ''),
                        article.get('realtorName', ''),
                        article.get('buildingName', ''),
                        ', '.join(article.get('tagList', [])),
                        article.get('articleFeatureDesc', ''),
                    ])

            if not data.get('isMoreData', False):
                break
            
            page += 1
        
        all_articles.sort(key=lambda x: self.parse_price(x[3]))  # Sort by price
        all_articles = self.remove_duplicate_lists(all_articles)
        
        return all_articles

    def save_data(self, all_articles):
        wb = Workbook()
        
        # Loop through each apartment's data and create a new sheet for each
        for apt_idx, data in self.apt_data.items():
            ws = wb.create_sheet(title=f"{data['apt_name']}")
            
            # Set column headers
            headers = [
                'Article Name', 'Real Estate Type', 'Trade Type', 'Price', 'Area Name',
                'Area2', 'Floor', 'Direction', 'Realtor Name', 'Building Name', 'Tags', 'Features'
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = header
            
            # Add article data
            apt_articles = [article for article in all_articles if str(article[0]) in data['apt_name']]  # Filter articles by area2

            for row_num, article in enumerate(apt_articles, 2):
                for col_num, value in enumerate(article, 1):
                    col_letter = get_column_letter(col_num)
                    ws[f'{col_letter}{row_num}'] = value
        
        # Remove the default sheet created with the workbook
        del wb['Sheet']

        # 특정 포맷으로 출력
        formatted_date = datetime.now().strftime("%Y-%m-%d")

        file_name = formatted_date + "_" + APT_LIST_OUTPUT
        # output 폴더 경로 추가
        output_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

        # output 폴더가 없으면 생성
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        # 파일을 output 폴더 안에 저장
        file_path = os.path.join(output_directory, file_name)
        wb.save(file_path)
        print(f"Data saved to {file_name}.")

        return file_path

    def send_email(self, subject, body, to_email, attachment_path):
        from_email = self.email_sender
        from_password = self.email_smtp_pw

        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        with open(attachment_path, "rb") as attachment:
            part = MIMEApplication(attachment.read(), Name=attachment_path)
            part['Content-Disposition'] = f'attachment; filename="{attachment_path}"'
            msg.attach(part)

        try:
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(from_email, from_password)
                server.sendmail(from_email, to_email, msg.as_string())
                print(f"Email sent to {to_email}.")
        except Exception as e:
            print(f"Failed to send email: {e}")
    
    def format_apt_data(self, apt_name, apt_data = list()):
        white_check = "✅"
        header = f'{white_check} {apt_name} {white_check}\n'
        # 필드별 최대 길이를 저장할 리스트
        max_lengths = [0] * 6  # 6은 `/`로 구분되는 필드의 개수

        # 각 필드의 최대 길이 계산
        for data in apt_data:
            fields = [data[3], data[5], data[6], data[7], data[9], data[11]]
            for i, field in enumerate(fields):
                max_lengths[i] = max(max_lengths[i], len(str(field)))

        formatted_str = ''
        # 데이터를 포맷팅하여 출력 문자열 생성
        for idx, data in enumerate(apt_data):
            fields = [data[3], data[5], data[6], data[9], data[7], data[11]]
            formatted_fields = [
                str(field).ljust(max_lengths[i]) for i, field in enumerate(fields)
            ]
            formatted_str += f'{str("📌").ljust(2)} {" | ".join(formatted_fields)}\n\n'
        
        formatted_str += '\n'

        return header + formatted_str

if __name__ == '__main__':
    print(os.path.join(os.getcwd(), APT_LIST_INPUT))
    scraper = RealEstateScraper(REAL_STATE_DATA_COOKIE, REAL_STATE_DATA_HEADER, os.path.join(os.path.dirname(os.path.abspath(__file__)), APT_LIST_INPUT))
    slack = SlackManager(SLACK_CHANNEL_ID, SLACK_OAUTH_TOKEN)
    # Fetch and save data for each apartment index
    all_articles = []
    for apt_idx, data in scraper.apt_data.items():
        apt_articles = scraper.fetch_data(apt_idx)
        all_articles.extend(apt_articles)

    file_path = scraper.save_data(all_articles)
    slack.upload_file(file_path)

    for apt_idx, apt_data in scraper.apt_data.items():
        my_data = []
        my_data.extend([data for data in all_articles 
                        if str(apt_data['apt_name']) == data[0] and
                        (((data[6].split("/")[0].isdigit() and int(data[6].split("/")[0]) >= LOWER_FLOOR_LIMIT) or 
                          ((data[6].split("/")[0] == "중") or (data[6].split("/")[0] == "고"))) and
                        (data[6].split("/")[0] != "저"))]) # 저층 제외
        message = scraper.format_apt_data(str(apt_data['apt_name']), my_data[:SLACK_LIST_LIMIT])

        print(message)
        slack.send_message(message)
    '''
    # Send the excel file as an email attachment
    scraper.send_email(
        subject="Real Estate Data for All Apartments",
        body="Please find the attached real estate data for all apartments.",
        to_email=EMAIL_RECEIVER,
        attachment_path=excel_file
    )
    '''
